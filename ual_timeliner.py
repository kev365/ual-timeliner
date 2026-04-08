#!/usr/bin/env python3

"""ual_timeliner: UAL Timeline Generator.
Build timeline-friendly data from Windows Server User Access Logging (UAL) databases.

By: BriMor Labs, extended by Kevin Stokes
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from ipaddress import AddressValueError, IPv4Address, IPv6Address
import os
from pathlib import Path
import shutil
import tempfile
from typing import Any, Iterable, Iterator, Mapping, Literal, Sequence
from uuid import UUID

import polars as pl
import pyesedb

# --- Constants & Configuration ---
# GUID_LOOKUP maps known RoleGUIDs to human-readable Role Names.
# These GUIDs identify specific Windows Server roles installed on the system.

GUID_LOOKUP: Mapping[str, str] = {
    "{10A9226F-50EE-49D8-A393-9A501D47CE04}": "File Server",
    "{4116A14D-3840-4F42-A67F-F2F9FF46EB4C}": "Windows Deployment Services",
    "{48EED6B2-9CDC-4358-B5A5-8DEA3B2F3F6A}": "DHCP Server",
    "{7CC4B071-292C-4732-97A1-CF9A7301195D}": "FAX Server",
    "{7FB09BD3-7FE6-435E-8348-7D8AEFB6CEA3}": "Print and Document Services",
    "{910CBAF9-B612-4782-A21F-F7C75105434A}": "BranchCache",
    "{952285D9-EDB7-4B6B-9D85-0C09E3DA0BBD}": "Remote Access",
    "{B4CDD739-089C-417E-878D-855F90081BE7}": "Active Directory Rights Management Service",
    "{BBD85B29-9DCC-4FD9-865D-3846DCBA75C7}": "Network Policy and Access Services",
    "{C23F1C6A-30A8-41B6-BBF7-F266563DFCD6}": "FTP Server",
    "{C50FCC83-BC8D-4DF5-8A3D-89D7F80F074B}": "Active Directory Certificate Services",
    "{D6256CF7-98FB-4EB4-AA18-303F1DA1F770}": "Web Server",
    "{D8DC1C8E-EA13-49CE-9A68-C9DCA8DB8B33}": "Windows Server Update Services",
    "{AD495FC3-0EAA-413D-BA7D-8B13FA7EC598}": "Active Directory Domain Services",
    "{BD7F7C0D-7C36-4721-AFA8-0BA700E26D9E}": "SQL Server Database Engine",
    "{DDE30B98-449E-4B93-84A6-EA86AF0B19FE}": "MSMQ",
    "{1479A8C1-9808-411E-9739-2D3C5923E86A}": "Remote Desktop Gateway",
    "{90E64AFA-70DB-4FEF-878B-7EB8C868F091}": "Remote Desktop Services",
    "{2414BC1B-1572-4CD9-9CA5-65166D8DEF3D}": "SQL Server Analysis Services",
    "{8CC0AC85-40F7-4886-9DAB-021519800418}": "Reporting Services",
    "{4AD13311-EC3B-447E-9056-14EDE9FA7052}": "Active Directory Lightweight Directory Services",
}

FILETIME_EPOCH = datetime(1601, 1, 1, tzinfo=timezone.utc)
IGNORED_NAMES = {"systemidentity.mdb"}
AnchorPreference = Literal[
    "insert_then_last",
    "last_then_insert",
    "last_only",
    "insert_only",
]

TIMELINE_FIELDS = [
    "timestamp",
    "timestamp_desc",
    "source_table",
    "authenticated_user",
    "ip_address",
    "host_name",
    "user",
    "access_count",
    "total_accesses",
    "role_name",
    "role_guid",
    "tenant_id",
    "client_name",
    "source_file",
]

TIMELINE_SCHEMA: dict[str, pl.PolarsDataType] = {
    "timestamp": pl.Datetime(time_zone="UTC"),
    "timestamp_desc": pl.Utf8,
    "source_table": pl.Utf8,
    "authenticated_user": pl.Utf8,
    "ip_address": pl.Utf8,
    "host_name": pl.Utf8,
    "user": pl.Utf8,
    "access_count": pl.Int64,
    "total_accesses": pl.Int64,
    "role_name": pl.Utf8,
    "role_guid": pl.Utf8,
    "tenant_id": pl.Utf8,
    "client_name": pl.Utf8,
    "source_file": pl.Utf8,
}

# --- Core Logic ---

@dataclass(frozen=True)
class TimelineEvent:
    """Container for a single timeline event row."""

    timestamp: datetime
    timestamp_description: str
    source_table: str
    source_file: Path
    role_guid: str | None = None
    role_name: str | None = None
    tenant_id: str | None = None
    client_name: str | None = None
    authenticated_user: str | None = None
    ip_address: str | None = None
    host_name: str | None = None
    user: str | None = None
    total_accesses: int | None = None
    access_count: int | None = None
    day_number: int | None = None

    def to_row(self) -> dict[str, Any]:
        """Return a dict suitable for building a DataFrame."""
        base: dict[str, Any] = {field: None for field in TIMELINE_FIELDS}
        base["timestamp"] = self.timestamp
        base["timestamp_desc"] = self.timestamp_description
        base["source_table"] = self.source_table
        base["authenticated_user"] = self.authenticated_user
        base["ip_address"] = self.ip_address
        base["host_name"] = self.host_name
        base["user"] = self.user
        base["total_accesses"] = self.total_accesses
        base["access_count"] = self.access_count
        base["role_name"] = self.role_name
        base["role_guid"] = self.role_guid
        base["tenant_id"] = self.tenant_id
        base["client_name"] = self.client_name
        base["source_file"] = str(self.source_file)
        return base


def build_timeline_from_directory(
    root: Path,
    anchor_preference: AnchorPreference = "insert_then_last",
    deduplicate: bool = True,
    full_output: bool = False,
    recursive: bool = False,
) -> pl.DataFrame:
    """
    Build a timeline DataFrame from every eligible .mdb in a directory.
    
    This function locates all .mdb files in the specified root directory (and subdirectories
    if recursive is True), parses them, and aggregates the results into a single Polars DataFrame.
    It handles file discovery and delegates the parsing logic to `build_timeline`.
    """
    paths = sorted(_find_mdb_files(root, recursive=recursive))
    return build_timeline(
        paths,
        anchor_preference=anchor_preference,
        deduplicate=deduplicate,
        full_output=full_output,
    )


def build_timeline(
    paths: Iterable[Path],
    anchor_preference: AnchorPreference = "insert_then_last",
    deduplicate: bool = True,
    full_output: bool = False,
) -> pl.DataFrame:
    """
    Build a timeline DataFrame from one or more .mdb files.

    Processes each file independently into a DataFrame, then concatenates
    the results. This avoids holding all raw events in memory simultaneously.

    Args:
        paths: A sequence of Path objects pointing to .mdb files.
        anchor_preference: Strategy for determining the year for Day### columns.
        deduplicate: If True, removes duplicate entries preferring 'Current.mdb'.
        full_output: If True, includes all extra columns and historical Day### data.

    Returns:
        A Polars DataFrame containing the processed timeline data.
    """
    frames: list[pl.DataFrame] = []
    for path in paths:
        print(f"[*] Processing: {path.resolve()}", file=sys.stderr)
        try:
            events = _read_mdb(
                path,
                anchor_preference=anchor_preference,
                full_output=full_output,
            )
        except Exception as e:
            error_str = str(e)
            if "pyesedb_file_open: unable to open file" in error_str:
                print(f"[*] Warning: Skipping {path.name} - Unable to open file (may be locked/dirty or invalid).", file=sys.stderr)
            else:
                print(f"[*] Warning: Skipping {path.name} - {e}", file=sys.stderr)
            continue
        if events:
            rows = [event.to_row() for event in events]
            frames.append(pl.from_dicts(rows, schema=TIMELINE_SCHEMA))
            del events, rows
    if not frames:
        df = pl.DataFrame(schema=TIMELINE_SCHEMA)
    else:
        df = pl.concat(frames)
        del frames
        if deduplicate:
            df = _deduplicate_timeline(df)
        df = df.sort(["timestamp", "timestamp_desc"])
    if not full_output:
        exclude_cols = ["role_guid", "client_name", "tenant_id", "access_count"]
        df = df.drop([c for c in exclude_cols if c in df.columns])
    return df


def _read_mdb(
    path: Path, anchor_preference: AnchorPreference, full_output: bool
) -> list[TimelineEvent]:
    """
    Extract events from a single ESE database.

    Opens the ESE database using pyesedb, retrieves the DNS, ROLE_ACCESS, and
    CLIENTS tables, and parses them into TimelineEvent objects.

    Args:
        path: Path to the .mdb file.
        anchor_preference: Strategy for year determination.
        full_output: Whether to extract dense historical data (Day### columns).

    Returns:
        A list of TimelineEvent objects found in the database.
    
    Raises:
        FileNotFoundError: If the path does not exist.
        pyesedb.error: If the database is corrupt or cannot be opened.
    """
    if path.name.lower() in IGNORED_NAMES:
        return []
    if not path.exists():
        msg = f"Database file not found: {path}"
        raise FileNotFoundError(msg)

    with _open_ese_db(path) as database:
        dns_table = _get_table(database, "DNS")
        role_table = _get_table(database, "ROLE_ACCESS")
        clients_table = _get_table(database, "CLIENTS")

        if role_table is None:
            raise LookupError("Table ROLE_ACCESS not found in database")
        if clients_table is None:
            raise LookupError("Table CLIENTS not found in database")

        events: list[TimelineEvent] = []
        if dns_table is not None:
            events.extend(_build_dns_events(dns_table, path))
        events.extend(_build_role_events(role_table, path))
        events.extend(
            _build_client_events(
                clients_table=clients_table,
                source_file=path,
                anchor_preference=anchor_preference,
                full_output=full_output,
            )
        )
        return events


@contextmanager
def _open_ese_db(path: Path) -> Iterator[pyesedb.file]:
    """
    Context manager to open an ESE DB, with fallback for dirty shutdown states.

    If pyesedb fails to open the file (common with dirty shutdowns), this manager
    attempts to create a temporary copy, patch the header state to 'Clean', and
    open that instead.
    """
    database = pyesedb.file()
    temp_path: Path | None = None

    # Attempt open — try direct first, then dirty-patch fallback.
    # This is separated from the yield to avoid catching caller exceptions.
    try:
        database.open(str(path))
    except Exception:
        temp_path = _create_clean_temp_copy(path)
        if not temp_path:
            raise
        database.open(str(temp_path))

    try:
        yield database
    finally:
        database.close()
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def _create_clean_temp_copy(path: Path) -> Path | None:
    """
    Create a temporary copy of the DB and patch its state to Clean Shutdown.
    
    Returns None if the file doesn't appear to be a dirty ESE DB or if IO fails.
    """
    try:
        # Read the state at offset 52 (0x34)
        with open(path, "rb") as f:
            header = f.read(64)
            
        if len(header) < 60:
            return None
            
        # 2 = Dirty Shutdown in ESE
        state = int.from_bytes(header[52:56], "little")
        if state != 2:
            return None
            
        # Create temp file with restricted permissions (owner-only read/write)
        fd, temp_path_str = tempfile.mkstemp(suffix=".mdb", prefix="ual_timeliner_")
        try:
            os.fchmod(fd, 0o600)
        except (AttributeError, OSError):
            pass  # fchmod unavailable on Windows; default perms are user-only
        os.close(fd)
        temp_path = Path(temp_path_str)
        
        # Copy and patch
        shutil.copy2(path, temp_path)
        with open(temp_path, "r+b") as f:
            f.seek(52)
            f.write((3).to_bytes(4, "little")) # 3 = Clean Shutdown
            
        print(f"[*] Info: Patched dirty shutdown state for {path.name}", file=sys.stderr)
        return temp_path
        
    except Exception:
        return None


def _get_table(database: pyesedb.file, name: str) -> pyesedb.table | None:
    """
    Return a table by name, or ``None`` if it does not exist.

    Iterates through all tables in the ESE database to find a match (case-insensitive).

    Args:
        database: An open pyesedb.file object.
        name: The name of the table to retrieve.

    Returns:
        The pyesedb.table object if found, otherwise ``None``.
    """
    for index in range(database.get_number_of_tables()):
        table = database.get_table(index)
        if table.get_name().upper() == name.upper():
            return table
    return None


def _build_dns_events(
    table: pyesedb.table, source_file: Path
) -> list[TimelineEvent]:
    """
    Parse DNS table into timeline events.

    The DNS table contains a mapping of IP addresses to Hostnames and the last time
    this mapping was observed.

    Args:
        table: The 'DNS' table from the ESE database.
        source_file: The path to the source .mdb file.

    Returns:
        List of TimelineEvent objects representing DNS sightings.
    """
    events: list[TimelineEvent] = []
    for index in range(table.get_number_of_records()):
        record = table.get_record(index)
        last_seen = _filetime_to_datetime(record.get_value_data(0))
        address = _decode_text(record.get_value_data(1))
        host_name = _decode_text(record.get_value_data(2))
        if last_seen is None:
            continue
        events.append(
            TimelineEvent(
                timestamp=last_seen,
                timestamp_description="LastSeen",
                source_table="DNS",
                source_file=source_file,
                ip_address=address,
                host_name=host_name,
            )
        )
    return events


def _build_role_events(
    table: pyesedb.table, source_file: Path
) -> list[TimelineEvent]:
    """
    Parse ROLE_ACCESS table into timeline events.

    The ROLE_ACCESS table tracks when specific Windows Roles (e.g., File Server,
    Active Directory) were first and last seen on the system.

    Args:
        table: The 'ROLE_ACCESS' table from the ESE database.
        source_file: The path to the source .mdb file.

    Returns:
        List of TimelineEvent objects representing Role first/last-seen times.
    """
    columns = _column_map(table)
    events: list[TimelineEvent] = []
    for index in range(table.get_number_of_records()):
        record = table.get_record(index)
        role_guid = _guid_to_str(record.get_value_data(columns.get("RoleGuid", 0)))
        role_name = GUID_LOOKUP.get(role_guid) if role_guid else None
        first_seen = _filetime_to_datetime(
            record.get_value_data(columns["FirstSeen"]) if "FirstSeen" in columns else record.get_value_data(1)
        )
        last_seen = _filetime_to_datetime(
            record.get_value_data(columns["LastSeen"]) if "LastSeen" in columns else None
        )
        if first_seen is not None:
            events.append(
                TimelineEvent(
                    timestamp=first_seen,
                    timestamp_description="FirstSeen",
                    source_table="ROLE_ACCESS",
                    source_file=source_file,
                    role_guid=role_guid,
                    role_name=role_name,
                )
            )
        if last_seen is not None:
            events.append(
                TimelineEvent(
                    timestamp=last_seen,
                    timestamp_description="LastSeen",
                    source_table="ROLE_ACCESS",
                    source_file=source_file,
                    role_guid=role_guid,
                    role_name=role_name,
                )
            )
    return events


def _build_client_events(
    clients_table: pyesedb.table,
    source_file: Path,
    anchor_preference: AnchorPreference,
    full_output: bool,
) -> list[TimelineEvent]:
    """
    Parse CLIENTS table into timeline events.

    The CLIENTS table is the primary source of user activity data. It contains
    records of users accessing specific roles from specific IP addresses.
    It includes 'InsertDate' (first access), 'LastAccess' (last access), and
    a series of 'Day###' columns tracking access counts for specific days of the year.

    Args:
        clients_table: The 'CLIENTS' table from the ESE database.
        source_file: The path to the source .mdb file.
        anchor_preference: Strategy for resolving the year for 'Day###' columns.
        full_output: If True, parses and includes 'Day###' historical access events.

    Returns:
        List of TimelineEvent objects representing user access activity.
    """
    events: list[TimelineEvent] = []
    columns = _column_map(clients_table)
    
    # Identify "Day###" columns which represent historical daily access counts.
    # We sort them by day number to process them sequentially.
    day_columns = sorted(
        ((name, idx) for name, idx in columns.items() if name.startswith("Day")),
        key=lambda pair: int(pair[0][3:]),
    )
    for index in range(clients_table.get_number_of_records()):
        record = clients_table.get_record(index)
        role_guid = _guid_to_str(record.get_value_data(columns["RoleGuid"]))
        role_name = GUID_LOOKUP.get(role_guid) if role_guid else None
        tenant_id = _guid_to_str(record.get_value_data(columns["TenantId"]))
        total_accesses = _safe_int(
            record.get_value_data_as_integer(columns["TotalAccesses"])
        )
        insert_date = _filetime_to_datetime(record.get_value_data(columns["InsertDate"]))
        last_access = _filetime_to_datetime(record.get_value_data(columns["LastAccess"]))
        raw_address_bytes = record.get_value_data(columns["Address"])
        _, ip_address = _convert_address(raw_address_bytes)
        authenticated_user = _decode_text(
            record.get_value_data(columns["AuthenticatedUserName"])
        )
        client_name = _decode_text(record.get_value_data(columns["ClientName"]))
        host_name = _host_from_authenticated_user(authenticated_user)
        user = _user_from_authenticated_user(authenticated_user)
        base_event_kwargs = {
            "source_table": "CLIENTS",
            "source_file": source_file,
            "role_guid": role_guid,
            "role_name": role_name,
            "tenant_id": tenant_id,
            "client_name": client_name,
            "authenticated_user": authenticated_user,
            "ip_address": ip_address,
            "host_name": host_name,
            "user": user,
            "total_accesses": total_accesses,
        }
        if insert_date:
            events.append(
                TimelineEvent(
                    timestamp=insert_date,
                    timestamp_description="InsertDate",
                    **base_event_kwargs,
                )
            )
        if last_access:
            events.append(
                TimelineEvent(
                    timestamp=last_access,
                    timestamp_description="LastAccess",
                    **base_event_kwargs,
                )
            )
        day_year = _day_year(insert_date, last_access, anchor_preference)
        if day_year is None or not full_output:
            continue
        for name, idx in day_columns:
            count = record.get_value_data_as_integer(idx)
            if count in (None, 0):
                continue
            day_number = int(name[3:])
            day_timestamp = datetime(day_year, 1, 1, tzinfo=timezone.utc) + timedelta(
                days=day_number - 1
            )
            events.append(
                TimelineEvent(
                    timestamp=day_timestamp,
                    timestamp_description=name,
                    access_count=int(count),
                    **base_event_kwargs,
                )
            )
    return events


def _find_mdb_files(root: Path, recursive: bool = False) -> Iterator[Path]:
    """
    Yield .mdb files under a directory, ignoring SystemIdentity.

    Scans the given root path. If it's a file, checks if it's a valid .mdb.
    If it's a directory, searches for .mdb files (recursively if requested).
    Explicitly ignores 'SystemIdentity.mdb' and files starting with '$I' (Recycle Bin metadata).

    Args:
        root: The root file or directory to search.
        recursive: If True, searches subdirectories.

    Yields:
        Path objects for each valid .mdb file found.
    """
    if root.is_file():
        if root.suffix.lower() == ".mdb" and root.name.lower() not in IGNORED_NAMES:
            yield root
        return
    
    resolved_root = root.resolve()
    pattern = "**/*.mdb" if recursive else "*.mdb"
    for path in root.glob(pattern):
        # Skip symlinks to avoid following links outside the target directory
        if path.is_symlink():
            continue
        # Ensure the resolved path is still under the root directory
        try:
            path.resolve().relative_to(resolved_root)
        except ValueError:
            continue
        # Ignore system metadata files (Recycle Bin $I, etc.) and known ignored names
        if path.name.lower() in IGNORED_NAMES or path.name.startswith("$I"):
            continue
        yield path


# --- Deduplication Logic ---

def _deduplicate_timeline(df: pl.DataFrame) -> pl.DataFrame:
    """
    Deduplicate rows, prioritizing data from 'Current.mdb' over historical backups.
    
    UAL data is often duplicated between the live 'Current.mdb' and the yearly
    archived GUID-named databases. This function removes exact duplicates based on
    key fields (timestamp, user, IP, etc.), ensuring that if a record exists in both,
    the version from 'Current.mdb' is kept (arbitrary but consistent choice,
    often the live DB is the 'primary' source until archival).
    """
    if "source_file" not in df.columns:
        return df
    
    # Specific fields requested for deduplication
    subset_cols = [
        "timestamp", 
        "timestamp_desc", 
        "source_table", 
        "authenticated_user", 
        "ip_address"
    ]
    
    # Ensure priority prefers Current.mdb (case-insensitive)
    priority = (
        pl.when(pl.col("source_file").str.to_lowercase().str.ends_with("current.mdb"))
        .then(0)
        .otherwise(1)
        .alias("_dedup_priority")
    )
    
    # Sort by the data fields + the priority to ensure the first unique row is the preferred one
    ordered = df.with_columns(priority).sort(subset_cols + ["_dedup_priority"])
    unique = ordered.unique(subset=subset_cols, keep="first")
    
    return unique.drop("_dedup_priority")


def _column_map(table: pyesedb.table) -> dict[str, int]:
    """
    Return a mapping of column name to index.

    Constructs a dictionary where keys are column names and values are their
    0-based indices in the ESE table. This avoids hardcoding indices and handles
    schema variations more robustly.

    Args:
        table: A pyesedb.table object.

    Returns:
        Dict mapping column name (str) to index (int).
    """
    return {
        table.get_column(idx).get_name(): idx
        for idx in range(table.get_number_of_columns())
    }


def _guid_to_str(value: bytes | str | None) -> str | None:
    """
    Convert a GUID to a normalized uppercase string with braces.
    
    Example: {12345678-ABCD-EF00-1234-56789ABCDEF0}
    
    Args:
        value: Raw byte sequence (16 bytes) or string representation of a GUID.

    Returns:
        Formatted GUID string or None if conversion fails.
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value.upper()
    try:
        guid = UUID(bytes_le=value)
    except (ValueError, AttributeError, TypeError):
        return None
    return f"{{{str(guid).upper()}}}"


def _decode_text(value: bytes | str | None) -> str | None:
    """Decode UTF-16LE values, returning None for empty strings."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip("\x00").strip()
        return cleaned or None
    decoded = bytes(value).decode("utf-16-le", errors="ignore").strip("\x00").strip()
    return decoded or None


def _host_from_authenticated_user(authenticated_user: str | None) -> str | None:
    """
    Extract hostname from domain\\hostname$ patterns.
    
    Computer accounts in Active Directory usually end with '$'. This function
    checks if the authenticated user field represents a machine account and
    extracts the hostname if so.

    Args:
        authenticated_user: The value from the 'AuthenticatedUserName' column.

    Returns:
        The extracted hostname (str) or None.
    """
    if authenticated_user is None or "$" not in authenticated_user:
        return None
    parts = authenticated_user.split("\\")
    candidate = parts[-1]
    if not candidate.endswith("$"):
        return None
    hostname = candidate[:-1].strip()
    return hostname or None


def _user_from_authenticated_user(authenticated_user: str | None) -> str | None:
    """
    Extract user from domain\\user patterns without trailing $.

    If the authenticated user is NOT a machine account (no trailing '$'),
    this function extracts the username part (after the last backslash).

    Args:
        authenticated_user: The value from the 'AuthenticatedUserName' column.

    Returns:
        The extracted username (str) or None.
    """
    if authenticated_user is None or "$" in authenticated_user:
        return None
    parts = authenticated_user.split("\\")
    candidate = parts[-1].strip()
    return candidate or None


def _filetime_to_datetime(value: bytes | int | None) -> datetime | None:
    """
    Convert Windows FILETIME to timezone-aware datetime in UTC.

    Windows FILETIME is a 64-bit integer representing the number of 100-nanosecond
    intervals since January 1, 1601 (UTC).

    Args:
        value: 64-bit integer or byte sequence.

    Returns:
        datetime object in UTC or None if the value is invalid/zero.
    """
    if value in (None, 0):
        return None
    if isinstance(value, bytes):
        int_value = int.from_bytes(value, "little", signed=False)
    else:
        int_value = int(value)
    microseconds = int_value // 10
    try:
        return FILETIME_EPOCH + timedelta(microseconds=microseconds)
    except OverflowError:
        return None


def _convert_address(raw: bytes | None) -> tuple[str | None, str | None]:
    """
    Convert raw bytes to hexadecimal string and normalized IP address.
    
    The 'Address' field in the CLIENTS table stores IP addresses as raw bytes.
    - 4 bytes: IPv4
    - 16 bytes: IPv6
    
    Returns:
        tuple: (Hex representation, Normalized IP string)
    """
    if raw is None:
        return None, None
    hex_value = raw.hex().upper()
    ip_value: str | None = None
    try:
        if len(raw) == 4:
            ip_value = str(IPv4Address(raw))
        elif len(raw) == 16:
            ip_value = str(IPv6Address(raw))
    except AddressValueError:
        ip_value = None
    return hex_value, ip_value


def _day_year(
    insert_date: datetime | None,
    last_access: datetime | None,
    anchor_preference: AnchorPreference,
) -> int | None:
    """
    Choose a year to anchor day columns.
    
    The 'Day###' columns in UAL data only indicate the day of the year (1-366),
    not the specific year. This function infers the correct year based on the
    record's 'InsertDate' and 'LastAccess' timestamps and the user's preference.

    Args:
        insert_date: The timestamp of the first access.
        last_access: The timestamp of the last access.
        anchor_preference: Strategy name (e.g., 'insert_then_last').

    Returns:
        The integer year to use for the Day### columns, or None if undetermined.
    """
    if anchor_preference == "insert_then_last":
        if insert_date is not None:
            return insert_date.year
        if last_access is not None:
            return last_access.year
    elif anchor_preference == "last_then_insert":
        if last_access is not None:
            return last_access.year
        if insert_date is not None:
            return insert_date.year
    elif anchor_preference == "last_only":
        if last_access is not None:
            return last_access.year
    elif anchor_preference == "insert_only":
        if insert_date is not None:
            return insert_date.year
    return None


def _safe_int(value: Any) -> int | None:
    """
    Coerce a value to int when possible.

    Args:
        value: Any value (string, bytes, etc.).

    Returns:
        Integer value or None if conversion fails.
    """
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None

# --- CLI & Output Handling ---

def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(
        description="ual_timeliner: Build a UTC timeline from Windows UAL ESE databases."
    )
    parser.add_argument(
        "path",
        type=Path,
        help="Directory containing UAL .mdb files (Current.mdb and GUID.mdb).",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output filename and location. If omitted, prints CSV to stdout.",
    )
    parser.add_argument(
        "-f",
        "--format",
        choices=["csv", "parquet", "xlsx", "sqlite", "k2t"],
        default="csv",
        help=(
            "Output format. Use k2t for Timesketch-compatible JSONL, "
            "xlsx for Excel, or sqlite for a SQLite database."
        ),
    )
    parser.add_argument(
        "--no-dedup",
        action="store_true",
        help="Disable deduplication (enabled by default).",
    )
    parser.add_argument(
        "--full-output",
        action="store_true",
        help="Include all output columns, parses Day### data, and no deduplication.",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="Recursively search for .mdb files.",
    )
    parser.add_argument(
        "--split-rows",
        type=int,
        default=0,
        help="Split output into multiple files every N rows (csv/k2t only).",
    )
    return parser.parse_args(argv)


def write_output(
    df: pl.DataFrame, output: Path | None, fmt: str, split_rows: int = 0
) -> None:
    """
    Write the DataFrame to disk or stdout.

    Supports multiple formats (CSV, Excel, SQLite, Parquet, K2T) and handles
    optional row splitting for CSV and K2T output.

    Args:
        df: The Polars DataFrame to write.
        output: Destination Path object (file). Can be None for stdout (CSV only).
        fmt: Output format string.
        split_rows: If > 0, splits output into chunks of N rows (CSV/K2T).

    Raises:
        ValueError: If output path is missing for file-based formats.
    """
    if output is not None and output.exists():
        stem = output.stem
        suffix = output.suffix
        counter = 1
        while output.exists():
            output = output.with_name(f"{stem}-{counter}{suffix}")
            counter += 1

    # Handle splitting for CSV and K2T
    if split_rows > 0 and output is not None and fmt in ("csv", "k2t"):
        total_rows = df.height
        if total_rows > split_rows:
            # Determine chunk count
            chunk_count = (total_rows + split_rows - 1) // split_rows
            base_stem = output.stem
            
            # Iterate and write each chunk as a separate file
            for i in range(chunk_count):
                start = i * split_rows
                length = min(split_rows, total_rows - start)
                chunk_df = df.slice(start, length)
                
                # Construct new filename: filename_partX.ext
                # We simply append _partXXX to the stem. The recursive call will handle
                # specific format extensions (like .json -> .jsonl for k2t).
                part_path = output.with_name(f"{base_stem}_part{i+1:03d}{output.suffix}")
                
                 # Recursively call write_output with split_rows=0 to write this specific chunk only.
                write_output(chunk_df, part_path, fmt, split_rows=0)
            return

    # Strip timezone offset for cleaner output in user-facing formats.
    if fmt in ("csv", "xlsx", "sqlite", "parquet") and "timestamp" in df.columns:
        df = df.with_columns(
            pl.col("timestamp").dt.replace_time_zone(None)
        ).rename({"timestamp": "timestamp (UTC)"})

    if fmt == "csv":
        if output is None:
            df.write_csv(sys.stdout.buffer)
            return
        output.parent.mkdir(parents=True, exist_ok=True)
        df.write_csv(output)
        return
    if output is None:
        msg = "Output path is required for parquet, xlsx, sqlite, and k2t formats."
        raise ValueError(msg)
    output.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "parquet":
        df.write_parquet(output)
        return
    if fmt == "xlsx":
        _write_xlsx(df, output)
        return
    if fmt == "sqlite":
        _write_sqlite(df, output)
        return
    if fmt == "k2t":
        if output.suffix.lower() == ".json":
            output = output.with_suffix(".jsonl")
        elif output.suffix.lower() != ".jsonl":
            output = output.with_name(f"{output.name}.jsonl")
        _write_k2t_jsonl(df, output)
        return
    msg = f"Unsupported format requested: {fmt}"
    raise ValueError(msg)


def _write_xlsx(df: pl.DataFrame, output: Path) -> None:
    """
    Write the timeline to XLSX with sheet splitting and auto-filter.

    Uses openpyxl write_only mode for streaming writes — rows are flushed to
    disk incrementally instead of building the full workbook in memory.

    Excel has a row limit (approx 1M). This function splits the DataFrame into
    multiple worksheets if it exceeds 900,000 rows.

    Args:
        df: The Polars DataFrame.
        output: The target .xlsx file path.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    max_rows_per_sheet = 900_000
    total_rows = df.height
    workbook = Workbook(write_only=True)
    sheet_count = max(1, (total_rows + max_rows_per_sheet - 1) // max_rows_per_sheet)
    headers = list(df.columns)
    end_col_letter = get_column_letter(len(headers))
    for sheet_index in range(sheet_count):
        sheet = workbook.create_sheet(title=f"timeline_{sheet_index + 1}")
        sheet.freeze_panes = "A2"
        sheet.append(headers)
        start = sheet_index * max_rows_per_sheet
        length = min(max_rows_per_sheet, total_rows - start)
        row_count = 0
        if length > 0:
            chunk = df.slice(start, length)
            for row in chunk.iter_rows():
                sheet.append([
                    value.isoformat() if isinstance(value, datetime) else value
                    for value in row
                ])
                row_count += 1
        sheet.auto_filter.ref = f"A1:{end_col_letter}{row_count + 1}"
    workbook.save(output)


def _write_sqlite(df: pl.DataFrame, output: Path) -> None:
    """
    Write the timeline to a SQLite database.

    Creates a table named 'timeline' and inserts all records.
    If the table exists, it appends data.

    Args:
        df: The Polars DataFrame.
        output: The target .sqlite file path.
    """
    connection = sqlite3.connect(output)
    try:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS timeline (
                "timestamp (UTC)" TEXT,
                timestamp_desc TEXT,
                source_table TEXT,
                authenticated_user TEXT,
                ip_address TEXT,
                host_name TEXT,
                user TEXT,
                access_count INTEGER,
                total_accesses INTEGER,
                role_name TEXT,
                role_guid TEXT,
                tenant_id TEXT,
                client_name TEXT,
                source_file TEXT
            )
            """
        )
        insert_sql = """
            INSERT INTO timeline (
                "timestamp (UTC)",
                timestamp_desc,
                source_table,
                authenticated_user,
                ip_address,
                host_name,
                user,
                access_count,
                total_accesses,
                role_name,
                role_guid,
                tenant_id,
                client_name,
                source_file
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        batch_size = 10_000
        batch: list[tuple[Any, ...]] = []
        for row in df.iter_rows(named=True):
            ts = row.get("timestamp (UTC)")
            ts_str = ts.isoformat() if isinstance(ts, datetime) else None
            batch.append(
                (
                    ts_str,
                    row.get("timestamp_desc"),
                    row.get("source_table"),
                    row.get("authenticated_user"),
                    row.get("ip_address"),
                    row.get("host_name"),
                    row.get("user"),
                    row.get("access_count"),
                    row.get("total_accesses"),
                    row.get("role_name"),
                    row.get("role_guid"),
                    row.get("tenant_id"),
                    row.get("client_name"),
                    row.get("source_file"),
                )
            )
            if len(batch) >= batch_size:
                connection.executemany(insert_sql, batch)
                batch.clear()
        if batch:
            connection.executemany(insert_sql, batch)
        connection.commit()
    finally:
        connection.close()


def _write_k2t_jsonl(df: pl.DataFrame, output: Path) -> None:
    """
    Write Timesketch-compatible JSONL (k2t).

    This format is typically converted to JSONL where each line is a JSON object
    containing a 'message', 'datetime', and 'timestamp_desc', along with other attributes.
    This is suitable for ingestion into Timesketch via OpenSearch.

    Args:
        df: The Polars DataFrame.
        output: The target file path (usually .jsonl).
    """
    required_fields = {"timestamp", "timestamp_desc"}
    columns = df.columns
    extras = [col for col in columns if col not in required_fields]
    message_fields = [col for col in extras if col != "source_file"]
    with output.open("w", encoding="utf-8") as handle:
        for row in df.iter_rows(named=True):
            ts = row.get("timestamp")
            if isinstance(ts, datetime):
                dt_value = ts.isoformat()
            else:
                dt_value = None
            payload: dict[str, Any] = {
                "message": _build_message(row, message_fields),
                "datetime": dt_value,
                "timestamp_desc": row.get("timestamp_desc"),
            }
            for key in extras:
                value = row.get(key)
                if isinstance(value, datetime):
                    payload[key] = value.isoformat()
                elif value is not None:
                    payload[key] = value
            handle.write(json.dumps(payload, ensure_ascii=False))
            handle.write("\n")


def _build_message(row: dict[str, Any], fields: list[str]) -> str:
    """Compose a message string from available fields."""
    parts: list[str] = []
    for key in fields:
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, datetime):
            rendered = value.isoformat()
        else:
            rendered = str(value)
        parts.append(f"{key}={rendered}")
    return ("\n| ").join(parts)


def main(argv: Sequence[str] | None = None) -> int:
    """Run the CLI."""
    args = parse_args(sys.argv[1:] if argv is None else argv)
    timeline = build_timeline_from_directory(
        root=args.path.resolve(),
        deduplicate=not args.no_dedup,
        full_output=args.full_output,
        recursive=args.recursive,
    )
    write_output(timeline, args.output, args.format, split_rows=args.split_rows)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
